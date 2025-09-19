# Fichier test_simple.py
import requests
from bs4 import BeautifulSoup
import time
from flask import Flask, request , redirect , render_template , send_file, url_for, request 
from openpyxl import Workbook
import uuid
import os
import jsonify
import traceback

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
    'AppleWebKit/537.36 (KHTML, like Gecko) '
    'Chrome/91.0.4472.124 Safari/537.36'
}


app = Flask(__name__)         

@app.route('/' , methods=['GET'])
def index():
    query = ''
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Produits'
    ws['B1'] = 'PRIX'
    ws['C1'] = 'Liens'
 
    
    if request.method == "GET":
        query = request.args.get('q', '').lower()
        session_id = str(uuid.uuid4())[:8]

        
    
    if query:
        
        search_type = request.args.get('search-type')
        category = request.args.get('category')
        max_price = int(request.args.get('max-price'))
        min_price = int(request.args.get('min-price'))
        
        name_list = []
        prod_price = []
        img_list = []
        link_list = []
        origin_list = []
        reduction_list = []
        date_list = []
        quant_list = []
        rate_list =[]
        inventaire = []


        if search_type == "fast":  
            if request.args.get('jumia'):
                url_j = f"https://www.jumia.ci/catalog/?q={query}&sort=lowest-price&price={min_price}-{max_price}" 
                response = requests.get(url_j, headers=headers)
                soup_j = BeautifulSoup(response.content, 'html.parser')
                containers = soup_j.find_all('article' , class_ = "prd _fb col c-prd")
                for container in containers :
                    name = container.find(class_ = 'name')
                    price = int(container.find(class_ = 'prc').text.split()[0].replace(',' , ''))
                    img = container.find('img' , class_ = 'img')['data-src']
                    link = f"https://jumia.ci{container.find('a', class_ = 'core')['href']}"
                    reduction = container.find('div' , class_ = 'bdg _dsct _sm')
                    rate = container.find('div' , class_ = 'rev')
                    if rate != None:
                        rate = float(container.find('div' , class_ = 'rev').find('div' , class_ = 'stars _s').text.split()[0])
                    else:
                        rate = "N/A"
                    response = requests.get(link , headers = headers)
                    soup = BeautifulSoup(response.content, 'html.parser')
                    quant = soup.find('p', class_ = '-df -i-ctr -fs12 -pbs -rd5')
                    if quant != None :
                        quant = quant.text.split()[0]
                    else:
                        quant = "N/A"

                    
                    date = 'N/A'
                    place = 'Livraison possible'
                    if reduction != None:
                        reduction = container.find('div' , class_ = 'bdg _dsct _sm').text.replace('%' , " %").split()[0]
                    else:
                        reduction = 0
                   
                    origin = "Jumia.ci"
                    if name.text.lower().find(query.lower()) == -1:
                        continue
                    name_list.append(name.text)
                    prod_price.append(price)
                    img_list.append(img)
                    link_list.append(link)
                    origin_list.append(origin)
                    reduction_list.append(reduction)
                    date_list.append(date)
                    quant_list.append(quant)
                    rate_list.append(rate)
                    
            
            
            if request.args.get('djokstore'):
                url_d = f'https://djokstore.ci/search?filter.v.availability=1&filter.v.price.gte={min_price}&filter.v.price.lte={max_price}&options%5Bprefix%5D=last&page=1&q={query}&sort_by=price-ascending'
                response = requests.get(url_d, headers=headers)
                soup_d = BeautifulSoup(response.content, 'html.parser')
                containers = soup_d.find_all('div' , class_ = "product-collection alo-border-radius")
                for container in containers:
                    name = container.find('div' , class_ = 'product-collection__content product-grid-style style_left').find('div' , class_ = 'product-collection__title').find('h2').find('a')
                    link = f"https://djokstore.ci{container.find('div' , class_ = 'product-collection__content product-grid-style style_left').find('div' , class_ = 'product-collection__title').find('h2').find('a')['href']}"
                    origin = 'Djokstore.ci'
                    place = 'f'
                    rate = 4
                    date = 'ddd'
                    try:
                        price = int(container.find('div' , class_= 'product-collection__content product-grid-style style_left').find('div' , class_= 'frm-price-color').find('div' , class_= 'product-collection__price').find('span' , class_= 'price price--sale').find('span' , class_= 'current').text.replace('FCFA' , '').replace(',' , ''))
                    except:
                        price = int(container.find('div' , class_= 'product-collection__content product-grid-style style_left').find('div' , class_= 'frm-price-color').find('div' , class_= 'product-collection__price').find('span' , class_= 'price').find('span').text.replace('FCFA' , '').replace(',' , ''))

                    try:
                        reduction =  container.find('div' , class_="product-collection__image product-image js-product-images-navigation js-product-images-hovered-end").find('div' , class_= 'product-image__overlay-top-left').find('div' , class_ ='label label--sale').text.replace('-' , '').replace('%' , '')
                    except:
                        reduction = 0


                    try:
                        img = container.find('div' , class_="product-collection__image product-image js-product-images-navigation js-product-images-hovered-end").find('div' , class_ = 'card__media').find('div', class_ = 'media secondary_image_hover').find('a', class_ = "d-block image_product cursor-default ratio").find('img' , class_ = 'motion-reduce default_media')['src']
                        
                    except:
                        pass
                    response = requests.get(link, headers=headers)
                    soup= BeautifulSoup(response.content, 'html.parser')
                    quant = int(soup.find('span', class_= 'qty').text)
                    if name.text.lower().find(query.lower()) == -1:
                                continue
                    name_list.append(name.text)
                    prod_price.append(price)
                    img_list.append(img)
                    link_list.append(link)
                    origin_list.append(origin)
                    reduction_list.append(reduction)
                    date_list.append(date)
                    quant_list.append(quant)
                    rate_list.append(rate)

           
            prod_count = len(name_list)
        elif search_type == "deep":
                i=0
                while True :
                    
                    i+=1 
                    if request.args.get('jumia'):
                        url = f"https://www.jumia.ci/catalog/?q={query}&sort=lowest-price&page={i}&price={min_price}-{max_price}"
                        response = requests.get(url, headers=headers)
                        soup = BeautifulSoup(response.content, 'html.parser')
                        containers = soup.find_all('article' , class_ = "prd _fb col c-prd")                
            
                        for container in containers :
                            name = container.find(class_ = 'name')
                            price = int(container.find(class_ = 'prc').text.split()[0].replace(',' , ''))
                            img = container.find('img' , class_ = 'img')['data-src']
                            link = f"https://jumia.ci{container.find('a', class_ = 'core')['href']}"
                            reduction = container.find('div' , class_ = 'bdg _dsct _sm')
                            rate = container.find('div' , class_ = 'rev')
                            if rate != None:
                                rate = float(container.find('div' , class_ = 'rev').find('div' , class_ = 'stars _s').text.split()[0])
                            else:
                                rate = "N/A"
                            response = requests.get(link , headers = headers)
                            soup = BeautifulSoup(response.content, 'html.parser')
                            quant = soup.find('p', class_ = '-df -i-ctr -fs12 -pbs -rd5')
                            if quant != None :
                                quant = quant.text.split()[0]
                            else:
                                quant = "N/A"

                           
                            date = 'N/A'
                            if reduction != None:
                                reduction = container.find('div' , class_ = 'bdg _dsct _sm').text.replace('%' , " %").split()[0]
                            else:
                                reduction = 0
                        
                            origin = "Jumia.ci"
                            if name.text.lower().find(query.lower()) == -1:
                                continue
                            name_list.append(name.text)
                            prod_price.append(price)
                            img_list.append(img)
                            link_list.append(link)
                            origin_list.append(origin)
                            reduction_list.append(reduction)
                            date_list.append(date)
                            quant_list.append(quant)
                            rate_list.append(rate)
                            
                        if containers == [] :
                                break       
                    if request.args.get('djokstore'):
                        url_d = f'https://djokstore.ci/search?filter.v.availability=1&filter.v.price.gte={min_price}&filter.v.price.lte={max_price}&options%5Bprefix%5D=last&page={i}&q={query}&sort_by=price-ascending'
                        response = requests.get(url_d, headers=headers)
                        soup_d = BeautifulSoup(response.content, 'html.parser')
                        containers_d = soup_d.find_all('div' , class_ = "product-collection alo-border-radius")
                        for container in containers_d:
                            name = container.find('div' , class_ = 'product-collection__content product-grid-style style_left').find('div' , class_ = 'product-collection__title').find('h2').find('a')
                            link = f"https://djokstore.ci{container.find('div' , class_ = 'product-collection__content product-grid-style style_left').find('div' , class_ = 'product-collection__title').find('h2').find('a')['href']}"
                            origin = 'Djokstore.ci'
                            place = 'f'
                            rate = 4
                            date = 'ddd'
                            try:
                                price = int(container.find('div' , class_= 'product-collection__content product-grid-style style_left').find('div' , class_= 'frm-price-color').find('div' , class_= 'product-collection__price').find('span' , class_= 'price price--sale').find('span' , class_= 'current').text.replace('FCFA' , '').replace(',' , ''))
                            except:
                                price = int(container.find('div' , class_= 'product-collection__content product-grid-style style_left').find('div' , class_= 'frm-price-color').find('div' , class_= 'product-collection__price').find('span' , class_= 'price').find('span').text.replace('FCFA' , '').replace(',' , ''))

                            try:
                                reduction =  container.find('div' , class_="product-collection__image product-image js-product-images-navigation js-product-images-hovered-end").find('div' , class_= 'product-image__overlay-top-left').find('div' , class_ ='label label--sale').text.replace('-' , '').replace('%' , '')
                            except:
                                reduction = 0


                            try:
                                img = container.find('div' , class_="product-collection__image product-image js-product-images-navigation js-product-images-hovered-end").find('div' , class_ = 'card__media').find('div', class_ = 'media secondary_image_hover').find('a', class_ = "d-block image_product cursor-default ratio").find('img' , class_ = 'motion-reduce default_media')['src']
                                
                            except:
                                pass
                            response = requests.get(link, headers=headers)
                            soup= BeautifulSoup(response.content, 'html.parser')
                            quant = int(soup.find('span', class_= 'qty').text)
                            if name.text.lower().find(query.lower()) == -1:
                                        continue
                            name_list.append(name.text)
                            prod_price.append(price)
                            img_list.append(img)
                            link_list.append(link)
                            origin_list.append(origin)
                            reduction_list.append(reduction)
                            date_list.append(date)
                            quant_list.append(quant)
                            rate_list.append(rate)
            
                        if containers_d == [] :
                            break   
                    
                    prod_count = len(name_list)
                
        
        for name,price,img,link,origin,reduction,date,quant,rate in zip(name_list , prod_price , img_list,link_list , origin_list , reduction_list, date_list ,quant_list,rate_list):
            inventaire.append({
                'name' : name , 
                'price' : price,
                'img' : img,
                'link' : link,
                'origin' : origin,
                'reduction': reduction,
                'date' : date,
                'quant' : quant,
                'rate': rate


            })
        inventaire_trié = sorted(inventaire, key=lambda x: x['price']) 
   
        return render_template('result.html',name_count= prod_count , name = name_list , prod_price = prod_price, inventaire = inventaire_trié , query = query , session_id = session_id)
    else:
        return render_template ('index.html')

@app.route('/compare')
def compare():
    return render_template('table.html')




@app.errorhandler(400)
def bad_request_error(error):
    return render_template('error.html', 
                         error_code=400,
                         error_title="Mauvaise Requête",
                         error_message="La requête est mal formée ou incomplète."), 400


@app.errorhandler(403)
def forbidden_error(error):
    return render_template('error.html',
                         error_code=403,
                         error_title="Accès Refusé",
                         error_message="Vous n'avez pas l'autorisation d'accéder à cette ressource."), 403


@app.errorhandler(404)
def not_found_error(error):
    return render_template('error.html',
                         error_code=404,
                         error_title="Page Non Trouvée",
                         error_message="La page que vous recherchez n'existe pas ou a été déplacée."), 404


@app.errorhandler(405)
def method_not_allowed_error(error):
    return render_template('error.html',
                         error_code=405,
                         error_title="Méthode Non Autorisée",
                         error_message="La méthode HTTP utilisée n'est pas autorisée pour cette URL."), 405


@app.errorhandler(500)
def internal_error(error):
    return render_template('error.html',
                         error_code=500,
                         error_title="Erreur Interne",
                         error_message="Une erreur interne du serveur s'est produite."), 500


@app.errorhandler(503)
def service_unavailable_error(error):
    return render_template('error.html',
                         error_code=503,
                         error_title="Service Indisponible",
                         error_message="Le service est temporairement indisponible."), 503



@app.errorhandler(Exception)
def handle_all_exceptions(error):
   
    app.logger.error(f"Unhandled exception: {str(error)}")
    app.logger.error(traceback.format_exc())
    
    if request.path.startswith('/api/'):
        return jsonify({
            'error': {
                'code': 500,
                'message': 'Internal server error',
                'details': str(error) if app.debug else 'An unexpected error occurred'
            }
        }), 500
    else:
        return render_template('error.html',
                             error_code=500,
                             error_title="Erreur Inattendue",
                             error_message="Une erreur s'est produite."), 500





if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port , debug=True)
