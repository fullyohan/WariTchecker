# Fichier test_simple.py
import requests
from bs4 import BeautifulSoup
import time
from flask import Flask, request , redirect , render_template , send_file, url_for 
from openpyxl import Workbook
from flask import Flask, redirect ,render_template , request
import uuid
import os
import jsonify
import traceback

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
    'AppleWebKit/537.36 (KHTML, like Gecko) '
    'Chrome/91.0.4472.124 Safari/537.36'
}


app = Flask(__name__)         

@app.route('/' , methods=['GET'])
def index():
    query = ''
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Produits'
    ws['B1'] = 'PRIX'
    ws['C1'] = 'Liens'
 
    
    if request.method == "GET":
        query = request.args.get('q', '').lower()
        session_id = str(uuid.uuid4())[:8]

        
    
    if query:
        
        search_type = request.args.get('search-type')
        max_price = int(request.args.get('max-price'))
        min_price = int(request.args.get('min-price'))
        excel_path = f"/home/elitebook/Desktop/price checker/Flask/static/download/result_{session_id}.xlsx"
        
        name_list = []
        prod_price = []
        img_list = []
        link_list = []
        origin_list = []
        inventaire = []


        if search_type == "fast":  
            url_j = f"https://www.jumia.ci/catalog/?q={query}&sort=lowest-price&price={min_price}-{max_price}" 
            response = requests.get(url_j, headers=headers)
            soup_j = BeautifulSoup(response.content, 'html.parser')
            containers = soup_j.find_all('article' , class_ = "prd _fb col c-prd")
            for container in containers :
                name = container.find(class_ = 'name')
                price = int(container.find(class_ = 'prc').text.split()[0].replace(',' , ''))
                img = container.find('img' , class_ = 'img')['data-src']
                link = container.find('a', class_ = "core")['href']
                origin = "Jumia.ci"
                if name.text.lower().find(query.lower()) == -1:
                    continue
                name_list.append(name.text)
                prod_price.append(price)
                img_list.append(img)
                link_list.append('https://jumia.ci' + link)
                origin_list.append(origin)
                ws.append([name.text , f"{price} FCFA",'https://jumia.ci' + link , origin] )
            
           

            url_c = f"https://ci.coinafrique.com/search?keyword={query}&sort_by=relevance&price_min={min_price}&price_max={max_price}&sort_by=price_asc"
            response = requests.get(url_c, headers=headers)
            soup = BeautifulSoup(response.content, 'html.parser')
            containers = soup.find_all('div' , class_ = "card ad__card round small hoverable undefined")
            for container in containers :
                name = container.find('div', class_= 'card-content ad__card-content').find('div').find('p' ,class_ = 'ad__card-description').find('a')
                price = int(container.find('p' , class_ = 'ad__card-price').text.replace("CFA", '').replace(" ", ''))
                link =container.find('p' , class_ = 'ad__card-price').find('a')['href']
                img = container.find('img' , class_ = 'ad__card-img')
                origin = "ci.CoinAfrique.com"
                if name.text.lower().find(query.lower()) == -1:
                    continue

                name_list.append(name.text)
                prod_price.append(price)
                link_list.append('https://ci.coinafrique.com' + link )
                origin_list.append(origin)
                ws.append([name.text , f"{price} FCFA" ,'https://ci.coinafrique.com' + link , origin] )
                img_list.append(img['src'])

            wb.save(excel_path)
            prod_count = len(name_list)
        elif search_type == "deep":
                i=0
                while True :
                    time.sleep(1)
                    i+=1 
                    url = f"https://www.jumia.ci/catalog/?q={query}&sort=lowest-price&page={i}&price={min_price}-{max_price}"
                    response = requests.get(url, headers=headers)
                    soup = BeautifulSoup(response.content, 'html.parser')
                    containers = soup.find_all('article' , class_ = "prd _fb col c-prd")                
        
                    for container in containers :
                        name = container.find(class_ = 'name')
                        price = int(container.find(class_ = 'prc').text.split()[0].replace(',' , ''))
                        img = container.find('img' , class_ = 'img')['data-src']
                        link = container.find('a', class_ = "core")['href']
                        origin = "Jumia.ci"
                        if name.text.lower().find(query.lower()) == -1:
                            continue
                        name_list.append(name.text)
                        prod_price.append(price)
                        img_list.append(img)
                        link_list.append('https://jumia.ci' + link)
                        origin_list.append(origin)
                        ws.append([name.text , f"{price} FCFA",'https://jumia.ci' + link , origin , f'Page{i}'] )
                        
                        
                

                    url_c = f"https://ci.coinafrique.com/search?keyword={query}&sort_by=relevance&price_min={min_price}&price_max={max_price}&sort_by=price_asc&page={i}"
                    response = requests.get(url_c, headers=headers)
                    soup = BeautifulSoup(response.content, 'html.parser')
                    containers_c = soup.find_all('div' , class_ = "card ad__card round small hoverable undefined")
                    for container in containers_c :
                        try:
                            name = container.find('div', class_= 'card-content ad__card-content').find('div').find('p' ,class_ = 'ad__card-description').find('a')
                            price = int(container.find('p' , class_ = 'ad__card-price').text.replace("CFA", '').replace(" ", ''))
                            link =container.find('p' , class_ = 'ad__card-price').find('a')['href']
                            img = container.find('img' , class_ = 'ad__card-img')
                            origin = "ci.CoinAfrique.com"
                            if name.text.lower().find(query.lower()) == -1:
                                continue
                            name_list.append(name.text)
                            prod_price.append(price) 
                            img_list.append(img['src'])
                            link_list.append('https://ci.coinafrique.com' + link )
                            origin_list.append(origin)
                            ws.append([name.text , f"{price} FCFA" ,'https://ci.coinafrique.com' + link , origin , f'Page{i}'] )
                        except:
                            pass
                   
                       



                    if containers == [] and containers_c == []:
                        break
                    
                    prod_count = len(name_list)
                
        
        for name,price,img,link,origin in zip(name_list , prod_price , img_list,link_list , origin_list):
            inventaire.append({
                'name' : name , 
                'price' : price,
                'img' : img,
                'link' : link,
                'origin' : origin
                
            })
        inventaire_trié = sorted(inventaire, key=lambda x: x['price']) 
        print(inventaire_trié)
        wb.save(excel_path)     
        
        return render_template('result.html',name_count= prod_count , name = name_list , prod_price = prod_price, inventaire = inventaire_trié , query = query , session_id = session_id)
    else:
        return render_template ('index.html' )








@app.route('/download/<session_id>')
def download(session_id):
    
    path = f"/home/elitebook/Desktop/price checker/Flask/static/download/result_{session_id}.xlsx"

    return send_file(path, as_attachment=True)



@app.route('/clear')
def clear():
    for file in os.listdir(path='/home/elitebook/Desktop/price checker/Flask/static/download') :
        if os.path.isfile("/home/elitebook/Desktop/price checker/Flask/static/download/" + file):
            if time.time() - os.path.getmtime("/home/elitebook/Desktop/price checker/Flask/static/download/" + file) > 3:
                os.remove("/home/elitebook/Desktop/price checker/Flask/static/download/" + file)
                

    return redirect(url_for('index'))

@app.errorhandler(400)
def bad_request_error(error):
    return render_template('error.html', 
                         error_code=400,
                         error_title="Mauvaise Requête",
                         error_message="La requête est mal formée ou incomplète."), 400


@app.errorhandler(403)
def forbidden_error(error):
    return render_template('error.html',
                         error_code=403,
                         error_title="Accès Refusé",
                         error_message="Vous n'avez pas l'autorisation d'accéder à cette ressource."), 403


@app.errorhandler(404)
def not_found_error(error):
    return render_template('error.html',
                         error_code=404,
                         error_title="Page Non Trouvée",
                         error_message="La page que vous recherchez n'existe pas ou a été déplacée."), 404


@app.errorhandler(405)
def method_not_allowed_error(error):
    return render_template('error.html',
                         error_code=405,
                         error_title="Méthode Non Autorisée",
                         error_message="La méthode HTTP utilisée n'est pas autorisée pour cette URL."), 405


@app.errorhandler(500)
def internal_error(error):
    return render_template('error.html',
                         error_code=500,
                         error_title="Erreur Interne",
                         error_message="Une erreur interne du serveur s'est produite."), 500


@app.errorhandler(503)
def service_unavailable_error(error):
    return render_template('error.html',
                         error_code=503,
                         error_title="Service Indisponible",
                         error_message="Le service est temporairement indisponible."), 503



@app.errorhandler(Exception)
def handle_all_exceptions(error):
   
    app.logger.error(f"Unhandled exception: {str(error)}")
    app.logger.error(traceback.format_exc())
    
    if request.path.startswith('/api/'):
        return jsonify({
            'error': {
                'code': 500,
                'message': 'Internal server error',
                'details': str(error) if app.debug else 'An unexpected error occurred'
            }
        }), 500
    else:
        return render_template('error.html',
                             error_code=500,
                             error_title="Erreur Inattendue",
                             error_message="Une erreur s'est produite."), 500





if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port , debug=True)
